#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
table_splitter.py - 病人用药表格拆分工具
将第二列的复合用药信息拆分为独立列，支持流式写入优化性能
"""

import sys
import re
import argparse
from collections import OrderedDict, defaultdict
from openpyxl import load_workbook
from openpyxl import Workbook
import warnings
warnings.filterwarnings('ignore')

# ========== 配置 ==========
DEBUG = True  # 打印详细处理过程

# 频度修饰词映射（不区分大小写）
FREQ_MAP = {
    'qd': 1, 'once': 1, 'onc': 1, 'oncem': 1, 'q3d': 1, 'w3d': 1,
    'bid': 2, 'tid': 3, 'qid': 4
}

# 用法词汇（需要移除）
USAGE_WORDS = ['口服', '皮下注射', '肌肉注射', '塞阴道', '外用', '吸入', '注射']

def clean_text(text):
    """清理文本：移除换行、多余空格、连续逗号（类似Excel CLEAN）"""
    if not text:
        return ""
    text = re.sub(r'[\r\n\t]', '', str(text))
    text = re.sub(r',+', ',', text)
    return text.strip()

def normalize_unit(unit_str, drug_name=""):
    """标准化单位，iu和mg互转处理"""
    if not unit_str:
        return ""
    unit_lower = unit_str.lower().strip()
    if 'iu' in unit_lower or unit_lower in ['i', 'u']:
        return 'iu'
    if 'mg' in unit_lower:
        return 'mg'
    if '片' in unit_lower:
        return '片'
    return unit_str

def parse_dose_amount(amount_str):
    """解析剂量数值和单位"""
    if not amount_str:
        return None, ""
    match = re.search(r'(\d+(?:\.\d+)?)', amount_str)
    if not match:
        return None, amount_str
    value = float(match.group(1))
    unit_part = amount_str[match.end():].strip()
    unit = normalize_unit(unit_part)
    return value, unit

def normalize_drug_name(name):
    """标准化药物名称：去除用法信息、日期、英文注释等"""
    if not name:
        return ""
    
    original_name = name
    
    # 移除日期格式 (2023-08-07) 或 (2023.08.07)
    name = re.sub(r'\([\d\-\.]+\)', '', name)
    
    # 移除英文注释括号内容如(HCG)，但保留中文括号内容
    name = re.sub(r'\([A-Za-z]+\)', '', name)
    
    # 移除用法词汇
    for uw in USAGE_WORDS:
        name = name.replace(uw, '')
    
    # 处理"皮下注射HMG" -> "HMG"
    match = re.search(r'(?:皮下注射|肌肉注射|注射)[,\s]*([A-Za-z0-9\u4e00-\u9fa5]+)', name)
    if match and len(match.group(1)) > 0:
        name = match.group(1)
    
    # 清理多余空格和标点
    name = re.sub(r'[,\s]+', ' ', name).strip()
    
    # 移除残留括号
    name = re.sub(r'[\(\)]', '', name)
    
    # 保护特定药品名
    if '达英-35' in original_name:
        name = '达英-35'
    
    return name if name else original_name

def parse_freq(freq_str):
    """解析频度字符串，返回天数（浮点数）"""
    if not freq_str:
        return 0.0
    freq_clean = freq_str.strip().lower()
    
    # 匹配 *数字天 或 ×数字天 格式
    match = re.search(r'[＊*×xX](\d+(?:\.\d+)?)\s*天', freq_clean)
    if match:
        return float(match.group(1))
    
    # 匹配 数字天 格式（无乘号）
    match2 = re.search(r'(\d+(?:\.\d+)?)\s*天', freq_clean)
    if match2:
        return float(match2.group(1))
    
    # 匹配频度修饰词
    for kw, mult in FREQ_MAP.items():
        if kw in freq_clean:
            num_match = re.search(r'(\d+(?:\.\d+)?)', freq_clean)
            if num_match:
                return float(num_match.group(1)) * mult
            return mult * 1.0
    
    return 0.0

def split_med_entries(text):
    """将第二列的文本拆分为基础格式列表"""
    if not text:
        return []
    
    # 移除日期格式避免干扰
    text = re.sub(r'\([\d\-\.]+\)', '', text)
    text_clean = clean_text(text)
    
    # 以"天"作为主要分隔符
    parts = re.split(r'(?<=天)', text_clean)
    raw_entries = [p.strip() for p in parts if p.strip()]
    
    # 如果没有找到"天"分隔，尝试用逗号
    if len(raw_entries) <= 1 and ',' in text_clean:
        raw_entries = [e.strip() for e in text_clean.split(',') if e.strip()]
    
    entries = []
    for raw in raw_entries:
        # 查找频度部分
        freq_match = re.search(r'[＊*×xX](\d+(?:\.\d+)?)\s*天', raw)
        if freq_match:
            days = float(freq_match.group(1))
            before_freq = raw[:freq_match.start()].strip()
            drug, dose_value, dose_unit = parse_drug_dose(before_freq)
            if drug:
                entries.append({
                    'drug': drug,
                    'dose_value': dose_value,
                    'dose_unit': dose_unit,
                    'days': days,
                    'raw': raw
                })
            else:
                if DEBUG:
                    print(f"    [警告] 无法解析药物: {raw[:50]}")
        else:
            # 尝试无乘号的格式（截断情况）
            simple_match = re.search(r'(\d+(?:\.\d+)?)\s*天$', raw)
            if simple_match:
                days = float(simple_match.group(1))
                before = raw[:simple_match.start()].strip()
                drug, dose_value, dose_unit = parse_drug_dose(before)
                if drug:
                    entries.append({
                        'drug': drug,
                        'dose_value': dose_value,
                        'dose_unit': dose_unit,
                        'days': days,
                        'raw': raw
                    })
            elif DEBUG and len(raw) > 3:
                print(f"    [警告] 无法识别频度，丢弃: {raw[:50]}")
    
    return entries

def parse_drug_dose(text):
    """从文本中解析药物名称和剂量"""
    if not text:
        return None, None, ""
    
    # 移除日期格式
    text = re.sub(r'\([\d\-\.]+\)', '', text)
    text = re.sub(r'[\d\-\.]+\s*$', '', text)
    
    # 查找剂量模式：数字+单位
    dose_pattern = r'(\d+(?:\.\d+)?)\s*([a-zA-Z\u4e00-\u9fa5]{0,5})?'
    match = re.search(dose_pattern, text)
    
    if not match:
        # 无剂量数值，整段作为药物名
        drug = normalize_drug_name(text.strip())
        return (drug, None, "") if drug and len(drug) > 1 else (None, None, "")
    
    dose_num = float(match.group(1))
    raw_unit = match.group(2) or ""
    unit = normalize_unit(raw_unit)
    
    # 药物名是剂量前面的部分
    drug_part = text[:match.start()].strip()
    if not drug_part:
        drug_part = text[match.end():].strip()
        if not drug_part:
            drug_part = "未知药物"
    
    drug = normalize_drug_name(drug_part)
    
    # 验证药物名有效性
    if drug and len(drug) > 1 and not re.match(r'^[\d\W]+$', drug):
        return (drug, dose_num, unit)
    
    return (None, None, "")

def merge_drug_key(drug, dose_value, dose_unit):
    """生成药品唯一标识键"""
    unit_str = dose_unit if dose_unit else ""
    if dose_value is None:
        return f"{drug}|无剂量|{unit_str}"
    # 剂量值保留2位小数用于比较
    return f"{drug}|{round(dose_value, 2)}|{unit_str}"

def process_sheet(filepath):
    """处理工作表，返回处理结果和药品字典"""
    print("\n" + "=" * 80)
    print("开始处理数据...")
    print("=" * 80)
    
    # 使用只读模式加载
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    
    all_drug_keys = OrderedDict()  # 药品键 -> 列索引
    rows_cache = []  # 存储每行处理结果
    all_med_stats = defaultdict(float)  # 统计所有药品总天数
    
    start_row = 3  # 跳过前2行描述
    current_row = start_row
    total_entries = 0
    total_discarded = 0
    reused_count = 0  # 复用前序行的次数
    
    for row in ws.iter_rows(min_row=start_row, values_only=False):
        cell_a = row[0]
        if cell_a.value is None or str(cell_a.value).strip() == '':
            print(f"\n行 {current_row}: 编号为空，停止处理后续行")
            break
        
        patient_id = str(cell_a.value).strip()
        cell_b = row[1] if len(row) > 1 else None
        raw_text = cell_b.value if cell_b and cell_b.value else ""
        raw_text = clean_text(str(raw_text))
        
        print(f"\n--- 处理行 {current_row} (编号: {patient_id}) ---")
        if len(raw_text) > 150:
            print(f"原始内容: {raw_text[:150]}...")
        else:
            print(f"原始内容: {raw_text}")
        
        # 拆分用药条目
        entries = split_med_entries(raw_text)
        total_entries += len(entries)
        print(f"识别到 {len(entries)} 个用药条目")
        
        # 合并相同(药物+剂量)的天数
        drug_days_map = defaultdict(float)
        for entry in entries:
            if entry['drug'] is None:
                total_discarded += 1
                continue
            
            key = merge_drug_key(entry['drug'], entry['dose_value'], entry['dose_unit'])
            drug_days_map[key] += entry['days']
            all_med_stats[key] += entry['days']
            
            dose_str = f"{entry['dose_value']}{entry['dose_unit']}" if entry['dose_value'] else "无剂量"
            print(f"  -> {entry['drug']} | {dose_str} | {entry['days']}天")
        
        # 记录复用情况
        new_drugs = [k for k in drug_days_map.keys() if k not in all_drug_keys]
        reused_drugs = [k for k in drug_days_map.keys() if k in all_drug_keys]
        if reused_drugs:
            reused_count += len(reused_drugs)
            print(f"  [复用] {len(reused_drugs)} 种药品已存在于前序行: {', '.join([d.split('|')[0] for d in reused_drugs[:3]])}")
        if new_drugs:
            print(f"  [新增] {len(new_drugs)} 种药品: {', '.join([d.split('|')[0] for d in new_drugs[:3]])}")
        
        process_msg = f"解析{len(entries)}条，合并{len(drug_days_map)}种"
        
        # 存储原始值
        original_values = [cell.value if cell.value is not None else "" for cell in row]
        
        # 注册新药品
        for drug_key in drug_days_map.keys():
            if drug_key not in all_drug_keys:
                all_drug_keys[drug_key] = len(all_drug_keys)
        
        rows_cache.append({
            'row_num': current_row,
            'patient_id': patient_id,
            'original_values': original_values,
            'process_msg': process_msg,
            'drugs': drug_days_map
        })
        
        current_row += 1
    
    wb.close()
    
    # 打印统计信息
    print("\n" + "=" * 80)
    print("处理完成统计:")
    print(f"  有效数据行数: {len(rows_cache)}")
    print(f"  总用药条目数: {total_entries}")
    print(f"  丢弃条目数: {total_discarded}")
    print(f"  药品种类数: {len(all_drug_keys)}")
    print(f"  药品复用次数: {reused_count}")
    
    print("\n药品清单（药物|剂量|单位）:")
    for i, key in enumerate(all_drug_keys.keys(), 1):
        total_days = all_med_stats.get(key, 0)
        print(f"  {i}. {key} (总天数: {total_days:.1f})")
    
    # 打印所有识别到的"药物名称 剂量"拼接
    print("\n所有识别到的药品组合（药物名称+剂量）:")
    for key in all_drug_keys.keys():
        parts = key.split('|')
        if len(parts) >= 2:
            print(f"  - {parts[0]} {parts[1]} {parts[2] if len(parts) > 2 else ''}")
    
    print("=" * 80)
    
    return rows_cache, all_drug_keys

def write_output(rows_cache, all_drug_keys, output_path, original_headers=None):
    """使用write_only模式高性能写入"""
    # 创建write_only模式的工作簿
    wb_out = Workbook(write_only=True)
    ws_out = wb_out.create_sheet(title="SplitResult")
    
    # 构建表头
    if original_headers and len(original_headers) > 0:
        headers = list(original_headers)
    else:
        max_cols = max((len(row['original_values']) for row in rows_cache), default=0)
        headers = [f"列{i+1}" for i in range(max_cols)]
    
    headers.append("处理过程信息")
    drug_headers = list(all_drug_keys.keys())
    headers.extend(drug_headers)
    
    # 写入表头
    ws_out.append(headers)
    
    # 逐行写入数据
    for row_data in rows_cache:
        new_row = []
        new_row.extend(row_data['original_values'])
        new_row.append(row_data['process_msg'])
        
        for drug_key in all_drug_keys.keys():
            days = row_data['drugs'].get(drug_key, "")
            if days != "":
                new_row.append(float(days))
            else:
                new_row.append(None)  # None在Excel中显示为空
        
        ws_out.append(new_row)
    
    # 保存文件
    wb_out.save(output_path)

def get_original_headers(filepath):
    """快速获取原始表头"""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=2, values_only=True):
        for val in row:
            if val:
                clean_val = re.sub(r'[\r\n\t]', '', str(val))
                headers.append(clean_val)
            else:
                headers.append("")
        break
    wb.close()
    
    # 去重保留顺序
    seen = set()
    unique_headers = []
    for h in headers:
        if h not in seen:
            seen.add(h)
            unique_headers.append(h)
    
    return unique_headers if unique_headers else None

def main():
    parser = argparse.ArgumentParser(description='病人用药表格拆分工具')
    parser.add_argument('input_file', help='输入Excel文件路径')
    parser.add_argument('output_file', nargs='?', help='输出Excel文件路径（可选）')
    args = parser.parse_args()
    
    input_path = args.input_file
    if args.output_file:
        output_path = args.output_file
    else:
        import os
        base = os.path.splitext(input_path)[0]
        output_path = f"{base}-o.xlsx"
    
    print(f"输入文件: {input_path}")
    print(f"输出文件: {output_path}")
    
    # 获取原始表头
    original_headers = get_original_headers(input_path)
    
    # 处理数据
    rows_cache, all_drug_keys = process_sheet(input_path)
    
    if not rows_cache:
        print("错误：没有有效数据行")
        return
    
    # 写入输出
    print("\n正在写入输出文件...")
    write_output(rows_cache, all_drug_keys, output_path, original_headers)
    print(f"完成！共写入 {len(rows_cache)} 行数据")
    print(f"结果已保存至: {output_path}")

if __name__ == "__main__":
    main()
